# -*- coding: utf-8 -*-
"""
Created on Mon Nov 18 09:26:30 2019

@author: GlaDo
"""
#import bug-free
from openpyxl import load_workbook
import operator

def Choose1(ListX,Student):#选题尝试第一轮
    if ListX[Student["LeveL1"]] == {}:
        ListX[Student["LeveL1"]] = Student
        Student["Grade"] = 0
def Choose2(ListX,Student):#选题尝试第二轮
    if ListX[Student["LeveL2"]] == {}:
        ListX[Student["LeveL2"]] = Student
        Student["Grade"] = 0
def Choose3(ListX,Student):#选题尝试第三轮
    if ListX[Student["LeveL3"]] == {}:
        ListX[Student["LeveL3"]] = Student
        Student["Grade"] = 0
def writeEXCEL(ws,ListX):#写入选题结果
    for i in range(len(ListX)):#写入选题结果List里的每一行
        if ListX[i] == {}:
            ws.append([])
        else:
            ws.append([i,ListX[i]["No"],ListX[i]["Name"]])
def writeEXCELF(ws,ListX):#写入选题结果
    for i in range(len(ListX)):#写入选题结果List里的每一行
        if ListX[i] == {}:
            ws.append([])
        else:
            ws.append([i,ListX[i]["No"],ListX[i]["Name"],ListX[i]["Grade"]])
                
#定义数据
Student = {"time":"","Name":"","No":"","Grade":"","Direction1":"","LeveL1":"","Direction2":"","LeveL2":"","Direction3":"","LeveL3":"","Flag":""}#一个学生的结构
ListStudent = []#学生集合
ListA = [{} for i in range(72)]#整车方向
ListB = [{} for i in range(58)]#电子方向
ListC = [{} for i in range(14)]#实验学方向
ListD = [{} for i in range(61)]#动力方向
ListE = [{} for i in range(31)]#车身方向
ListF = [{} for i in range(23)]#新能源方向
ListG = [{} for i in range(26)]#营销方向
Student=dict()
ListA.append(Student)#额外添加一个元素，因为list第一个元素编号为0
Student=dict()
ListB.append(Student)#额外添加一个元素，因为list第一个元素编号为0
Student=dict()
ListC.append(Student)#额外添加一个元素，因为list第一个元素编号为0
Student=dict()
ListD.append(Student)#额外添加一个元素，因为list第一个元素编号为0
Student=dict()
ListE.append(Student)#额外添加一个元素，因为list第一个元素编号为0
Student=dict()
ListF.append(Student)#额外添加一个元素，因为list第一个元素编号为0
Student=dict()
ListG.append(Student)#额外添加一个元素，因为list第一个元素编号为0

#读取学生列表
wb=load_workbook('D://test3.xlsx')
#获取sheet：
table = wb.get_sheet_by_name('20191125')   #通过表名获取  
#读取每一行学生信息 不读取首行 行数：table.max_row
for i in range(1,table.max_row):
    Student=dict()
    Student["time"] = table.cell(row = i,column = 1).value
    Student["Name"] = table.cell(row = i,column = 2).value
    Student["No"] = table.cell(row = i,column = 3).value
    Student["Grade"] = table.cell(row = i,column = 4).value
    Student["Direction1"] = table.cell(row = i,column = 5).value
    Student["LeveL1"] = table.cell(row = i,column = 6).value
    Student["Direction2"] = table.cell(row = i,column = 7).value
    Student["LeveL2"] = table.cell(row = i,column = 8).value
    Student["Direction3"] = table.cell(row = i,column = 9).value
    Student["LeveL3"] = table.cell(row = i,column = 10).value
    Student["Flag"] = table.cell(row = i,column = 11).value
    if Student["Flag"] == 1:
        Student["Grade"]=Student["Grade"]+10#对于保研和指定课题学生 绩点+10保证一志愿选到
    ListStudent.append(Student)
del ListStudent[0]#删除首行
   
#第一志愿选题         
ListStudent.sort(key=operator.itemgetter("Grade"),reverse = True)#以绩点对学生进行降序排序
for i in range(len(ListStudent)):#进行第一轮选题，选中的学生绩点清零
    if ListStudent[i]["Grade"]>0:
        if ListStudent[i]["Direction1"]=="A":
            Choose1(ListA,ListStudent[i])
        elif  ListStudent[i]["Direction1"]=="B":
            Choose1(ListB,ListStudent[i])
        elif  ListStudent[i]["Direction1"]=="C":
            Choose1(ListC,ListStudent[i])
        elif  ListStudent[i]["Direction1"]=="D":
            Choose1(ListD,ListStudent[i])
        elif  ListStudent[i]["Direction1"]=="E":
            Choose1(ListE,ListStudent[i])
        elif  ListStudent[i]["Direction1"]=="F":
            Choose1(ListF,ListStudent[i])
        elif  ListStudent[i]["Direction1"]=="G":
            Choose1(ListG,ListStudent[i])

#第二志愿选题 
ListStudent.sort(key=operator.itemgetter("Grade"),reverse = True)#以绩点对学生进行降序排序
for i in range(len(ListStudent)):#进行第二轮选题，选中的学生绩点清零
    if ListStudent[i]["Grade"]>0:
        if ListStudent[i]["Direction2"]=="A":
            Choose2(ListA,ListStudent[i])
        elif  ListStudent[i]["Direction2"]=="B":
            Choose2(ListB,ListStudent[i])
        elif  ListStudent[i]["Direction2"]=="C":
            Choose2(ListC,ListStudent[i])
        elif  ListStudent[i]["Direction2"]=="D":
            Choose2(ListD,ListStudent[i])
        elif  ListStudent[i]["Direction2"]=="E":
            Choose2(ListE,ListStudent[i])
        elif  ListStudent[i]["Direction2"]=="F":
            Choose2(ListF,ListStudent[i])
        elif  ListStudent[i]["Direction2"]=="G":
            Choose2(ListG,ListStudent[i])            

#第三志愿选题 
ListStudent.sort(key=operator.itemgetter("Grade"),reverse = True)#以绩点对学生进行降序排序
for i in range(len(ListStudent)):#进行第三轮选题，选中的学生绩点清零
    if ListStudent[i]["Grade"]>0:
        if ListStudent[i]["Direction3"]=="A":
            Choose3(ListA,ListStudent[i])
        elif  ListStudent[i]["Direction3"]=="B":
            Choose3(ListB,ListStudent[i])
        elif  ListStudent[i]["Direction3"]=="C":
            Choose3(ListC,ListStudent[i])
        elif  ListStudent[i]["Direction3"]=="D":
            Choose3(ListD,ListStudent[i])
        elif  ListStudent[i]["Direction3"]=="E":
            Choose3(ListE,ListStudent[i])
        elif  ListStudent[i]["Direction3"]=="F":
            Choose3(ListF,ListStudent[i])
        elif  ListStudent[i]["Direction3"]=="G":
            Choose3(ListG,ListStudent[i])            
                        
ListStudent.sort(key=operator.itemgetter("Grade"),reverse = True)#以绩点对学生进行降序排序 未查看未选中同学
#输出选题结果
wsA = wb.create_sheet("整车") #写入整车结果
writeEXCEL(wsA,ListA)
wsB = wb.create_sheet("电子") #写入电子结果
writeEXCEL(wsB,ListB)
wsC = wb.create_sheet("实验学") #写入实验学结果
writeEXCEL(wsC,ListC)
wsD = wb.create_sheet("动力") #写入动力结果
writeEXCEL(wsD,ListD)
wsE = wb.create_sheet("车身") #写入车声结果
writeEXCEL(wsE,ListE)
wsF = wb.create_sheet("新能源") #写入新能源结果
writeEXCEL(wsF,ListF)
wsG = wb.create_sheet("营销") #写入营销结果   
writeEXCEL(wsG,ListG)     
wsG = wb.create_sheet("Fail") #写入营销结果   
writeEXCELF(wsG,ListStudent)     
wb.save('D://test5.xlsx')             